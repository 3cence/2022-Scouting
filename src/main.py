from game import Game
import kivy
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.widget import Widget
from kivy.uix.popup  import Popup
from kivy.lang.builder import Builder
from kivy.properties import ObjectProperty
from kivy.uix.screenmanager import Screen, ScreenManager
from openpyxl import Workbook, worksheet, workbook
from openpyxl.reader.excel import load_workbook
import os

game = Game()

class MatchScreen (Screen):
    teamNum = ObjectProperty(None)
    matchNum = ObjectProperty(None)

    autonHighLbl = ObjectProperty(None)
    autonLowLbl = ObjectProperty(None)
    teleopHighLbl = ObjectProperty(None)
    teleopLowLbl = ObjectProperty(None)

    resetBtn = ObjectProperty(None)

    redTeamBtn = ObjectProperty(None)
    blueTeamBtn = ObjectProperty(None)

    resetAttemptsGiven = 5
    resetAttempts = resetAttemptsGiven

    def resetMatch(self, bypass=False):
        postmatchScreen = self.manager.get_screen("postmatch")
        if self.resetAttempts <= 0 or bypass:
            self.resetAttempts = self.resetAttemptsGiven
            self.resetBtn.text = "Reset"
            postmatchScreen.resetData()
            self.resetData()
        else:
            self.resetBtn.text = str(self.resetAttempts)
            self.resetAttempts -= 1

    def resetData(self):
        game.team = ""
        self.redTeamBtn.disabled = False
        self.blueTeamBtn.disabled = False
        self.teamNum.text = ""
        game.autonHigh = 0
        game.autonLow = 0
        game.teleopHigh = 0
        game.teleopLow = 0
        self.autonHighLbl.text = str(game.autonHigh)
        self.autonLowLbl.text = str(game.autonLow)
        self.teleopHighLbl.text = str(game.teleopHigh)
        self.teleopLowLbl.text = str(game.teleopLow)

    def teamSelect(self, team):
        if team == "red":
            game.team = "red"
            self.redTeamBtn.disabled = True
            self.blueTeamBtn.disabled = False
        elif team == "blue":
            game.team = "blue"
            self.redTeamBtn.disabled = False
            self.blueTeamBtn.disabled = True

    def addGoal(self, goal):
        if goal == "ah":
            game.autonHigh += 1
            self.autonHighLbl.text = str(game.autonHigh)
        elif goal == "al":
            game.autonLow += 1
            self.autonLowLbl.text = str(game.autonLow)
        elif goal == "th":
            game.teleopHigh += 1
            self.teleopHighLbl.text = str(game.teleopHigh)
        elif goal == "tl":
            game.teleopLow += 1
            self.teleopLowLbl.text = str(game.teleopLow)

    def removeGoal(self, goal):
        if goal == "ah":
            game.autonHigh -= 1
            if game.autonHigh < 0:
                game.autonHigh = 0
            self.autonHighLbl.text = str(game.autonHigh)
        elif goal == "al":
            game.autonLow -= 1
            if game.autonLow < 0:
                game.autonLow = 0
            self.autonLowLbl.text = str(game.autonLow)
        elif goal == "th":
            game.teleopHigh -= 1
            if game.teleopHigh < 0:
                game.teleopHigh = 0
            self.teleopHighLbl.text = str(game.teleopHigh)
        elif goal == "tl":
            game.teleopLow -= 1
            if game.teleopLow < 0:
                game.teleopLow = 0
            self.teleopLowLbl.text = str(game.teleopLow)


class PostmatchScreen (Screen):

    notesBox = ObjectProperty(None)
    climberSpin = ObjectProperty(None)
    barSpin = ObjectProperty(None)
    sucessSpin = ObjectProperty(None)
    scouterName = ObjectProperty(None)

    def resetData(self):
        self.notesBox.text = ""
        self.climberSpin.text = "No"
        self.barSpin.text = "None"
        self.sucessSpin.text = "No"

    def matchEnd(self):
        androidPath = "/storage/emulated/0/Download/"
        windowsPath = "./sheets/"
        path = androidPath
        if not os.path.isdir(path):
            os.makedirs(path)
        path += "2022data.xlsx"
        matchScreen = self.manager.get_screen("match")
        print("Team #:", matchScreen.teamNum.text, "Match:", matchScreen.matchNum.text, "Name:", self.scouterName.text, "Ah:", game.autonHigh, "Al:", game.autonLow, "Th:", game.teleopHigh, "Tl:", game.teleopLow, "Team:", game.team, "Climber:", self.climberSpin.text, "Attempt:", self.barSpin.text, "Sucess:", self.sucessSpin.text, "Notes:", self.notesBox.text)
        try:
            wb = load_workbook(path)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.append(["Team #", "Match #", "Team", "Auton High", "Auton Low", "Teleop High", "Teleop Low", "Climber", "Attempt", "Sucess", "Notes"])
        ws.append([matchScreen.teamNum.text, matchScreen.matchNum.text, game.team, game.autonHigh, game.autonLow, game.teleopHigh, game.teleopLow, self.climberSpin.text, self.barSpin.text, self.sucessSpin.text, self.notesBox.text])
        wb.save(path)
        matchScreen.matchNum.text = str(int(matchScreen.matchNum.text) + 1)
        matchScreen.resetMatch(True)


class ScreenManager (ScreenManager):
    pass

kv = Builder.load_file("MatchScreen.kv")

class MainWindow (App):
    def build(self):
        return kv

if __name__ == "__main__":
    MainWindow().run()